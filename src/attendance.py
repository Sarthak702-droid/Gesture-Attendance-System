import csv
import os
from datetime import datetime

CSV_PATH = os.path.join("data", "attendance.csv")
XLSX_PATH = os.path.join("data", "attendance.xlsx")
COOLDOWN_SECONDS = 300  # 5 minutes cooldown to prevent double logging

def sync_to_excel():
    """
    Reads the attendance CSV file and syncs it to a beautifully formatted Excel (.xlsx) file.
    Uses pandas and openpyxl to auto-fit columns.
    """
    if not os.path.exists(CSV_PATH) or os.stat(CSV_PATH).st_size == 0:
        return
        
    try:
        import pandas as pd
        
        # Read the CSV
        df = pd.read_csv(CSV_PATH)
        
        # Write to Excel
        with pd.ExcelWriter(XLSX_PATH, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Attendance_Records")
            
            # Access workbook and worksheet to format columns
            workbook = writer.book
            worksheet = writer.sheets["Attendance_Records"]
            
            # Format header cells (make them bold and add light blue background)
            from openpyxl.styles import Font, PatternFill, Alignment
            
            header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid") # Dark Blue
            center_alignment = Alignment(horizontal="center", vertical="center")
            
            # Format column headers
            for col_idx in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=1, column=col_idx)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                
            # Align data columns
            for row in worksheet.iter_rows(min_row=2, max_row=len(df) + 1, min_col=1, max_col=len(df.columns)):
                for cell in row:
                    # Center align status, date, time and timestamp columns
                    if cell.column in [2, 3, 4, 5, 6, 7]:
                        cell.alignment = center_alignment
            
            # Auto-fit column widths to prevent truncation
            for col in worksheet.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                worksheet.column_dimensions[col_letter].width = max(max_len + 4, 12)
                
    except ImportError:
        print("[WARNING] pandas or openpyxl not installed. Skipping Excel (.xlsx) sync.")
    except Exception as e:
        print(f"[WARNING] Failed to sync to Excel: {e}")

def mark_attendance(name, status, location, evidence_path):
    """
    Appends a new attendance record (including location & evidence snapshot) to the CSV.
    Prevents duplicate entries within the cooldown window (5 minutes).
    Auto-syncs the logs to a formatted Excel file.
    """
    # Ensure data directory exists
    os.makedirs(os.path.dirname(CSV_PATH), exist_ok=True)
    
    now = datetime.now()
    date_str = now.strftime("%Y-%m-%d")
    time_str = now.strftime("%H:%M:%S")
    current_ts = int(now.timestamp())
    
    file_exists = os.path.exists(CSV_PATH)
    
    # Check for duplicate entry within cooldown period
    if file_exists:
        try:
            with open(CSV_PATH, "r", newline="", encoding="utf-8") as f:
                reader = csv.reader(f)
                rows = list(reader)
                
                # Check rows from most recent to oldest
                for row in reversed(rows[1:]):  # Skip header row
                    if len(row) >= 7:
                        row_name, row_status, _, _, _, _, row_ts_str = row[0], row[1], row[2], row[3], row[4], row[5], row[6]
                        try:
                            row_ts = int(row_ts_str)
                            if row_name.strip().lower() == name.strip().lower() and row_status.strip().upper() == status.strip().upper():
                                time_diff = current_ts - row_ts
                                if time_diff < COOLDOWN_SECONDS:
                                    remaining = COOLDOWN_SECONDS - time_diff
                                    rem_min = remaining // 60
                                    rem_sec = remaining % 60
                                    return False, f"Already logged {status} for {name}. Cooldown: {rem_min}m {rem_sec}s remaining."
                        except ValueError:
                            continue
        except Exception as e:
            print(f"[WARNING] Could not read attendance CSV for duplicate checking: {e}")
            
    # Append the new attendance record
    try:
        write_header = not file_exists or os.stat(CSV_PATH).st_size == 0
        with open(CSV_PATH, "a", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            if write_header:
                writer.writerow(["Name", "Status", "Date", "Time", "Location", "Evidence_Path", "Timestamp"])
            writer.writerow([
                name.strip(), 
                status.strip().upper(), 
                date_str, 
                time_str, 
                location.strip(), 
                evidence_path.strip(), 
                current_ts
            ])
            
        # Sync to formatted Excel sheet
        sync_to_excel()
        
        return True, f"Successfully logged {status} for {name}!"
    except Exception as e:
        return False, f"Error writing to attendance file: {e}"
